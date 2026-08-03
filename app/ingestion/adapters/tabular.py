import csv
import json
from io import BytesIO, StringIO
from typing import Any

from openpyxl import load_workbook


class CsvAdapter:
    content_types = frozenset({"text/csv", "application/vnd.ms-excel"})

    def parse(self, content: bytes) -> list[dict[str, str]]:
        text = content.decode("utf-8-sig")
        return list(csv.DictReader(StringIO(text)))

    def normalize(self, parsed: list[dict[str, str]], metadata: dict[str, Any]):
        return parsed


class XlsxAdapter:
    content_types = frozenset(
        {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    )

    def parse(self, content: bytes) -> list[dict[str, Any]]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.active
        if worksheet is None:
            workbook.close()
            raise ValueError("Workbook has no active worksheet")
        values = worksheet.iter_rows(values_only=True)
        headers = next(values, None)
        if not headers:
            raise ValueError("Workbook is empty")
        names = [str(value).strip() if value is not None else "" for value in headers]
        if not all(names):
            raise ValueError("Workbook contains an empty column header")
        rows = [
            dict(zip(names, row, strict=True))
            for row in values
            if any(value is not None for value in row)
        ]
        workbook.close()
        return rows

    def normalize(self, parsed: list[dict[str, Any]], metadata: dict[str, Any]):
        return parsed


class JsonAdapter:
    content_types = frozenset({"application/json"})

    def parse(self, content: bytes) -> Any:
        return json.loads(content)

    def normalize(self, parsed: Any, metadata: dict[str, Any]):
        return parsed
